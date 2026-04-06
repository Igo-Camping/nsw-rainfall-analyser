from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"D:\Weather App Folder\Imports\Master Costing Tool - Stormwater.xlsm")
OUTPUT_PATH = Path(__file__).with_name("costing_lookup.json")


def clean(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def number_or_none(value):
    if value in (None, "", "-"):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def main() -> None:
    wb = load_workbook(WORKBOOK_PATH, data_only=False, keep_vba=True)
    wb_values = load_workbook(WORKBOOK_PATH, data_only=True, keep_vba=True)

    table_ws = wb["_Tables"]
    table_values_ws = wb_values["_Tables"]
    rows = []
    for r in range(2, table_ws.max_row + 1):
        asset_class = clean(table_ws.cell(r, 1).value)
        category = clean(table_ws.cell(r, 2).value)
        item = clean(table_ws.cell(r, 3).value)
        component = clean(table_ws.cell(r, 4).value)
        source = clean(table_ws.cell(r, 5).value)
        type_name = clean(table_ws.cell(r, 6).value)
        size = clean(table_ws.cell(r, 7).value)
        unit = clean(table_ws.cell(r, 8).value)
        unit_rate = table_values_ws.cell(r, 9).value
        modifier = clean(table_ws.cell(r, 10).value)

        if not asset_class or not item:
            continue

        rows.append({
            "assetClass": asset_class,
            "category": category,
            "item": item,
            "component": component,
            "source": source,
            "type": type_name,
            "size": size,
            "unit": unit,
            "unitRate": number_or_none(unit_rate),
            "modifier": modifier,
        })

    mod_ws = wb["IPART Modifiers"]
    modifiers = []
    for r in range(1, mod_ws.max_row + 1):
        context = clean(mod_ws.cell(r, 6).value)
        level = clean(mod_ws.cell(r, 7).value) or clean(mod_ws.cell(r, 11).value)
        factor = number_or_none(mod_ws.cell(r, 12).value)
        if not level or factor is None:
            continue
        modifiers.append({
            "context": context,
            "level": level,
            "factor": factor,
        })

    asset_classes = sorted({row["assetClass"] for row in rows if row["assetClass"]})
    categories = sorted({row["category"] for row in rows if row["category"]})
    items = sorted({row["item"] for row in rows if row["item"]})

    payload = {
        "meta": {
            "sourceWorkbook": str(WORKBOOK_PATH),
            "rowCount": len(rows),
            "assetClassCount": len(asset_classes),
            "categoryCount": len(categories),
            "itemCount": len(items),
        },
        "modifiers": modifiers,
        "assetClasses": asset_classes,
        "rows": rows,
    }

    OUTPUT_PATH.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH} with {len(rows)} rows and {len(modifiers)} modifiers")


if __name__ == "__main__":
    main()
