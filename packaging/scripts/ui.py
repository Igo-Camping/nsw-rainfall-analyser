# app.py

import streamlit as st
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import zipfile

from cost_engine import (
    # FIX: import ASSET_ID_COL from cost_engine rather than redefining it
    ASSET_ID_COL,
    ASSET_DIAM_COL,
    ASSET_LEN_COL,
    CONDITION_COL,
    split_streams,
    load_relining_rates,
    load_reconstruction_rates,
    get_relining_rate,
    get_reconstruction_rate,
    calc_reconstruction_cost,
    DEFAULT_PIPE_CLASS,
    split_into_value_packages,
    split_into_value_packages_twop,
    split_into_packages_by_count,
    assign_spatial_clusters,
    PROXIMITY_OPTIONS,
    X_MID_COL,
    Y_MID_COL,
    topup_packages,
    get_size_group,
    get_package_group_label,
    summarise_pipe_packages,
    resolve_suburb_column,
)

# ---------------------------------------------------------
# PAGE CONFIG + THEME
# ---------------------------------------------------------

st.set_page_config(
    page_title="Stormwater Packaging Tool",
    layout="wide",
    initial_sidebar_state="expanded",
    page_icon="🟢",
)

st.markdown("""
<style>
html, body, [class*="css"] { color: #0A7A3B !important; }
:root {
    --primary-color: #0A7A3B !important;
    --secondary-color: #0A7A3B !important;
    --text-color: #0A7A3B !important;
}
.stButton>button {
    background-color: #0A7A3B !important;
    color: white !important;
    border-radius: 6px !important;
    border: none !important;
}
.stButton>button:hover { background-color: #0C8F45 !important; }
[data-testid="stMetricValue"] { color: #0A7A3B !important; }
[data-testid="stMetricDelta"] { color: #0A7A3B !important; }
.stTabs [data-baseweb="tab"] { color: #0A7A3B !important; }
.stTabs [aria-selected="true"] {
    border-bottom: 3px solid #0A7A3B !important;
    color: #0A7A3B !important;
}
div[data-baseweb="select"] > div { border-color: #0A7A3B !important; }
div[role="radiogroup"] > div { color: #0A7A3B !important; }
h1, h2, h3, h4, h5, h6 { color: #0A7A3B !important; }
[data-testid="stDataFrame"] th { color: #0A7A3B !important; }
.stAlert > div { color: #0A7A3B !important; }



/* Constrain selectboxes and number inputs to fit their content */
div[data-baseweb="select"],
div[data-baseweb="select"] > div {
    min-width: 120px !important;
    max-width: 400px !important;
    width: fit-content !important;
}
div[data-testid="stNumberInput"] > div {
    max-width: 180px !important;
}
</style>
""", unsafe_allow_html=True)

# ---------------------------------------------------------
# FIXED PATHS
# ---------------------------------------------------------

import sys as _sys
import os as _os

def _get_assets_path():
    if getattr(_sys, "frozen", False):
        p = _os.path.join(_sys._MEIPASS, "data", "assets_with_coords.csv")
        if _os.path.isfile(p):
            return p
        return _os.path.join(_sys._MEIPASS, "data", "assets.csv")
    here = _os.path.dirname(_os.path.abspath(__file__))
    for filename in ["assets_with_coords.csv", "assets.csv"]:
        candidate = _os.path.join(here, "..", "data", filename)
        if _os.path.isfile(candidate):
            return _os.path.normpath(candidate)
    return _os.path.join(here, "..", "data", "assets_with_coords.csv")

ASSETS_PATH = _get_assets_path()

# ---------------------------------------------------------
# UTILITIES
# ---------------------------------------------------------

# Columns to keep in downloaded package CSVs — everything else is stripped
OUTPUT_KEEP_COLS = [
    "package_id",
    "Asset",
    "SWP_Pipe Diameter_mm",
    "Spatial Length_m",
    "SW_Condition",
    "SW LGA 20% H1-H6",
    "SWP_ Pipe Material",
    "SWP_ Pipe Class",
    "SW_Observation Metho",
    "SW_Upstream Node",
    "SW_Downstream Node",
    # Costing columns (added at runtime)
    "rate_per_m",
    "pipe_cost",
    "Pipe Rate ($/m)",
    "Pipe Cost ($)",
    "Trench Volume (m3)",
    "Waste Weight (t)",
    "Excavation Cost ($)",
    "Backfill Cost ($)",
    "Demolition Cost ($)",
    "Waste Disposal Cost ($)",
    "Total Cost ($)",
    "Pipe_Start_Address",
    "Asset Suburb",
    "Number of Pipes",
]


def get_contractors(rates: pd.DataFrame) -> list[str]:
    """Extract unique sorted contractor names from the rate table.
    Only includes contractors that have at least one non-zero Unit Rate
    against a core pricing item (excludes establishment/modifier-only rows
    and mean/average price rows)."""
    EXCLUDE_NAMES = ["mean", "average", "median", "stormwater mean"]
    # Items that are modifiers/establishment only — not actual work rates
    EXCLUDE_ITEMS = ["establishment", "modifier", "traffic control"]

    for col in ["Component", "Origin of Average/Mean Price"]:
        if rates is None or rates.empty or col not in rates.columns:
            continue

        # Filter to rows with a real unit rate
        r = rates.copy()
        if "Unit Rate" in r.columns:
            r["_rate_num"] = pd.to_numeric(r["Unit Rate"], errors="coerce").fillna(0)
            # Exclude modifier/establishment items
            if "Item" in r.columns:
                item_mask = ~r["Item"].astype(str).str.lower().str.contains(
                    "|".join(EXCLUDE_ITEMS), na=False
                )
                r = r[item_mask & (r["_rate_num"] > 0)]
            else:
                r = r[r["_rate_num"] > 0]

        all_vals = r[col].dropna().astype(str).str.strip()
        contractors = [
            v for v in all_vals.unique()
            if not any(excl in v.lower() for excl in EXCLUDE_NAMES)
        ]
        return sorted(contractors)
    return []

def df_to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8")


def create_zip_from_packages(
    df: pd.DataFrame,
    suburb_col: str | None = None,
    diam_col: str | None = None,
) -> io.BytesIO:
    buffer = io.BytesIO()
    csv_filenames = []
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as z:
        for pkg_id in sorted(df["package_id"].unique()):
            pkg_df = df[df["package_id"] == pkg_id]
            if pkg_df.empty:
                continue

            first = pkg_df.iloc[0]
            parts = [str(pkg_id)]

            if suburb_col and suburb_col in pkg_df.columns:
                suburb_val = str(first[suburb_col]).strip()
                if suburb_val:
                    parts.append(suburb_val.replace(" ", "_"))

            if diam_col and diam_col in pkg_df.columns:
                try:
                    diam_val = str(int(float(first[diam_col]))).strip()
                except (ValueError, TypeError):
                    diam_val = str(first[diam_col]).strip()
                if diam_val:
                    parts.append(f"D{diam_val}")

            # Add asset ID after package ID for single-pipe reconstruction packages only
            if ASSET_ID_COL in pkg_df.columns and len(pkg_df) == 1 and str(pkg_id).startswith("REC_"):
                asset_id = str(pkg_df.iloc[0][ASSET_ID_COL]).replace(".0", "").strip()
                if asset_id:
                    parts.insert(1, asset_id)

            filename = "_".join(parts) + ".csv"
            csv_filenames.append(filename)

            # For reconstruction packages, rename breakdown columns to readable headers
            out_df = pkg_df.copy()
            if str(pkg_id).startswith("REC_"):
                # Add a pipe-only cost column (rate × length) before renaming
                if "rate_per_m" in out_df.columns and ASSET_LEN_COL in out_df.columns:
                    out_df["_pipe_cost_only"] = out_df["rate_per_m"] * out_df[ASSET_LEN_COL]
                rec_rename = {
                    "_trench_vol": "Trench Volume (m3)",
                    "_waste_t": "Waste Weight (t)",
                    "_exc_cost": "Excavation Cost ($)",
                    "_bkf_cost": "Backfill Cost ($)",
                    "_dem_cost": "Demolition Cost ($)",
                    "_waste_cost": "Waste Disposal Cost ($)",
                    "rate_per_m": "Pipe Rate ($/m)",
                    "_pipe_cost_only": "Pipe Cost ($)",
                    "pipe_cost": "Total Cost ($)",
                }
                out_df = out_df.rename(columns={k: v for k, v in rec_rename.items() if k in out_df.columns})
            else:
                # Drop internal breakdown columns from non-rec packages
                drop_cols = [c for c in out_df.columns if c.startswith("_")]
                out_df = out_df.drop(columns=drop_cols, errors="ignore")

            # Strip columns not in the whitelist — only keep what's explicitly listed
            keep = [c for c in OUTPUT_KEEP_COLS if c in out_df.columns]
            out_df = out_df[keep]
            # Format pipe_cost as currency
            cost_col = "Total Cost ($)" if str(pkg_id).startswith("REC_") else "pipe_cost"
            if cost_col in out_df.columns:
                out_df[cost_col] = pd.to_numeric(out_df[cost_col], errors="coerce").apply(
                    lambda x: f"${x:,.2f}" if pd.notna(x) else ""
                )

            # Add totals row
            totals = {col: "" for col in out_df.columns}
            totals["package_id"] = "Totals"
            if ASSET_LEN_COL in out_df.columns:
                totals[ASSET_LEN_COL] = round(
                    pd.to_numeric(out_df[ASSET_LEN_COL], errors="coerce").sum(), 2
                )
            if cost_col in out_df.columns:
                raw_cost = pd.to_numeric(
                    out_df[cost_col].astype(str).str.replace(r"[$,]", "", regex=True),
                    errors="coerce"
                ).sum()
                totals[cost_col] = f"${raw_cost:,.2f}"
            if ASSET_ID_COL in out_df.columns:
                totals["Number of Pipes"] = len(out_df)
                out_df["Number of Pipes"] = ""

            out_df = pd.concat([out_df, pd.DataFrame([totals])], ignore_index=True)
            z.writestr(filename, out_df.to_csv(index=False).encode("utf-8"))

        # Build summary text file
        lines = []
        lines.append("PACKAGE SUMMARY")
        lines.append("=" * 60)
        for pkg_id in sorted(df["package_id"].unique()):
            pkg_df = df[df["package_id"] == pkg_id]
            if pkg_df.empty:
                continue

            first = pkg_df.iloc[0]
            lines.append(f"")
            lines.append(f"Package:    {pkg_id}")

            if suburb_col and suburb_col in pkg_df.columns:
                suburbs = ", ".join(sorted(pkg_df[suburb_col].astype(str).unique()))
                lines.append(f"Suburb:     {suburbs}")

            if diam_col and diam_col in pkg_df.columns:
                try:
                    diams = ", ".join(str(int(float(d))) for d in sorted(pkg_df[diam_col].dropna().unique()))
                except (ValueError, TypeError):
                    diams = ", ".join(str(d) for d in sorted(pkg_df[diam_col].astype(str).unique()))
                lines.append(f"Diameters:  {diams}mm")

            pipe_count = len(pkg_df)
            lines.append(f"Pipes:      {pipe_count}")

            if ASSET_LEN_COL in pkg_df.columns:
                total_len = pkg_df[ASSET_LEN_COL].sum()
                lines.append(f"Length:     {total_len:,.1f}m")

            if "pipe_cost" in pkg_df.columns:
                total_cost = pkg_df["pipe_cost"].sum()
                lines.append(f"Cost:       ${total_cost:,.0f}")

            if ASSET_ID_COL in pkg_df.columns:
                asset_ids = (
                    pkg_df[ASSET_ID_COL]
                    .dropna()
                    .astype(str)
                    .str.replace(r"\.0$", "", regex=True)
                    .tolist()
                )
                lines.append(f"Assets:     {', '.join(asset_ids)}")

            lines.append("-" * 60)

        summary_text = "\n".join(lines)
        z.writestr("package_summary.txt", summary_text.encode("utf-8"))

        # Build simple file list from collected filenames
        filelist_lines = ["PACKAGE FILE LIST", "=" * 60, ""]
        for name in sorted(csv_filenames):
            filelist_lines.append(name)
        z.writestr("file_list.txt", "\n".join(filelist_lines).encode("utf-8"))

        # ── Package summary as .xlsx ─────────────────────────────────────────
        wb_summary = openpyxl.Workbook()
        ws = wb_summary.active
        ws.title = "Package Summary"

        GREEN = "0A7A3B"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor=GREEN)
        headers = ["Package ID", "Suburb", "Diameters (mm)", "Pipes", "Length (m)", "Cost ($)", "Asset IDs"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, pkg_id in enumerate(sorted(df["package_id"].unique()), 2):
            pkg_df = df[df["package_id"] == pkg_id]
            if pkg_df.empty:
                continue
            suburb = ", ".join(sorted(pkg_df[suburb_col].astype(str).unique())) if suburb_col and suburb_col in pkg_df.columns else ""
            try:
                diams = ", ".join(str(int(float(d))) for d in sorted(pkg_df[diam_col].dropna().unique())) if diam_col and diam_col in pkg_df.columns else ""
            except (ValueError, TypeError):
                diams = ""
            pipe_count = len(pkg_df)
            total_len = round(pkg_df[ASSET_LEN_COL].sum(), 1) if ASSET_LEN_COL in pkg_df.columns else ""
            total_cost = round(pkg_df["pipe_cost"].sum(), 0) if "pipe_cost" in pkg_df.columns else ""
            asset_ids = ", ".join(pkg_df[ASSET_ID_COL].dropna().astype(str).str.replace(r"\.0$", "", regex=True).tolist()) if ASSET_ID_COL in pkg_df.columns else ""
            ws.append([pkg_id, suburb, diams, pipe_count, total_len, total_cost, asset_ids])

        # Auto-width columns
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        summary_xlsx = io.BytesIO()
        wb_summary.save(summary_xlsx)
        z.writestr("package_summary.xlsx", summary_xlsx.getvalue())

        # ── File list as .xlsx ───────────────────────────────────────────────
        wb_files = openpyxl.Workbook()
        ws2 = wb_files.active
        ws2.title = "File List"
        cell = ws2.cell(row=1, column=1, value="Filename")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=GREEN)
        for row_idx, name in enumerate(sorted(csv_filenames), 2):
            ws2.cell(row=row_idx, column=1, value=name)
        ws2.column_dimensions["A"].width = 80

        filelist_xlsx = io.BytesIO()
        wb_files.save(filelist_xlsx)
        z.writestr("file_list.xlsx", filelist_xlsx.getvalue())

    buffer.seek(0)
    return buffer


def render_costing_subtab(
    state_key: str,
    rates,
    rate_fn,
    label: str,
):
    """
    Renders the Package Costing subtab for a given stream.

    Parameters
    ----------
    state_key : session state prefix, e.g. "rel"
    rates     : rate table dataframe
    rate_fn   : get_relining_rate or get_reconstruction_rate
    label     : human-readable stream label, e.g. "Relining"
    """
    st.subheader(f"{label} Package Costing")

    packaged = st.session_state.get(f"{state_key}_packaged")

    if packaged is None or packaged.empty:
        st.info("Generate packages first in the Package Generation tab.")
        return

    pkg_list = sorted(packaged["package_id"].unique())
    selected_pkg = st.selectbox("Select package", pkg_list, key=f"{state_key}_cost_pkg_select")

    pkg_df = packaged[packaged["package_id"] == selected_pkg].copy()

    col1, col2, _ = st.columns([1, 1, 3])
    with col1:
        pkg_cost_mode = st.selectbox(
            "Cost mode",
            ["median", "lowest", "contractor"],
            key=f"{state_key}_pkg_cost_mode",
        )
    with col2:
        pkg_contractor = None
        if pkg_cost_mode == "contractor":
            contractors = get_contractors(rates)
            pkg_contractor = st.selectbox(
                "Contractor",
                contractors if contractors else ["No contractors found"],
                key=f"{state_key}_pkg_contractor",
            )

    if state_key == "rec":
        cost_results = pkg_df.apply(
            lambda r: calc_reconstruction_cost(
                diameter_mm=r.get(ASSET_DIAM_COL),
                length_m=r.get(ASSET_LEN_COL),
                rates=rates,
                mode="vendor" if pkg_cost_mode == "contractor" else pkg_cost_mode,
                vendor=pkg_contractor if pkg_cost_mode == "contractor" else None,
                pipe_class=DEFAULT_PIPE_CLASS,
            ),
            axis=1,
        )
        pkg_df["rate_per_m"] = cost_results.apply(lambda x: x["pipe_rate"] if x else None)
        pkg_df["pipe_cost"] = cost_results.apply(lambda x: x["total_cost"] if x else None)
        pkg_df["_exc_cost"] = cost_results.apply(lambda x: x["excavation_cost"] if x else None)
        pkg_df["_bkf_cost"] = cost_results.apply(lambda x: x["backfill_cost"] if x else None)
        pkg_df["_dem_cost"] = cost_results.apply(lambda x: x["demolition_cost"] if x else None)
        pkg_df["_waste_cost"] = cost_results.apply(lambda x: x["waste_cost"] if x else None)
        pkg_df["_trench_vol"] = cost_results.apply(lambda x: x["trench_volume_m3"] if x else None)
        pkg_df["_waste_t"] = cost_results.apply(lambda x: x["waste_weight_t"] if x else None)
    else:
        pkg_df["rate_per_m"] = pkg_df.apply(
            lambda row: rate_fn(
                row.get(ASSET_DIAM_COL),
                row.get(ASSET_LEN_COL),
                rates,
                mode="vendor" if pkg_cost_mode == "contractor" else pkg_cost_mode,
                vendor=pkg_contractor if pkg_cost_mode == "contractor" else None,
            ),
            axis=1,
        )
        pkg_df["pipe_cost"] = pkg_df["rate_per_m"] * pkg_df[ASSET_LEN_COL]

    # Warn about uncosted pipes
    uncosted_pkg = pkg_df[pkg_df["pipe_cost"].isna()].copy()
    if not uncosted_pkg.empty and state_key != "rel":
        st.warning(f"{len(uncosted_pkg)} pipe(s) had no matching rate and have been excluded from cost totals.")
        _suburb = resolve_suburb_column(pkg_df)
        uncosted_cols = [c for c in [ASSET_LEN_COL, _suburb, ASSET_DIAM_COL, ASSET_ID_COL] if c and c in uncosted_pkg.columns]
        st.dataframe(uncosted_pkg[uncosted_cols].reset_index(drop=True).rename(lambda x: x+1), use_container_width=True)

    st.subheader(f"Package {selected_pkg} Summary")
    k1, k2, k3 = st.columns(3)
    k1.metric("Pipes", len(pkg_df))
    k2.metric("Length (m)", f"{pkg_df[ASSET_LEN_COL].sum():,.1f}")
    k3.metric("Cost ($)", f"${pkg_df['pipe_cost'].sum():,.0f}")

    st.subheader("Pipes in this package")
    if state_key == "rec":
        display_cols = [c for c in [ASSET_ID_COL, ASSET_DIAM_COL, ASSET_LEN_COL,
                                     "rate_per_m", "_trench_vol", "_waste_t",
                                     "_exc_cost", "_bkf_cost", "_dem_cost", "_waste_cost", "pipe_cost"]
                        if c in pkg_df.columns]
        rename_map = {
            "rate_per_m": "Pipe Rate ($/m)",
            "_trench_vol": "Trench Vol (m³)",
            "_waste_t": "Waste (t)",
            "_exc_cost": "Excavation ($)",
            "_bkf_cost": "Backfill ($)",
            "_dem_cost": "Demolition ($)",
            "_waste_cost": "Waste Disposal ($)",
            "pipe_cost": "Total Cost ($)",
        }
    else:
        display_cols = [c for c in [ASSET_ID_COL, ASSET_DIAM_COL, ASSET_LEN_COL, "rate_per_m", "pipe_cost"]
                        if c in pkg_df.columns]
        rename_map = {"rate_per_m": "Rate ($/m)", "pipe_cost": "Total Cost ($)"}

    display_df = pkg_df[display_cols].rename(columns=rename_map)
    st.dataframe(display_df.reset_index(drop=True).rename(lambda x: x+1), use_container_width=True)

    st.download_button(
        label=f"Download package {selected_pkg} (CSV)",
        data=df_to_csv_bytes(pkg_df),
        file_name=f"{selected_pkg}.csv",
        mime="text/csv",
        key=f"{state_key}_cost_download",
    )


def render_packaging_tab(
    stream_df: pd.DataFrame,
    rates,
    rate_fn,
    prefix: str,
    state_key: str,
    empty_msg: str,
    zip_filename: str,
    tab_label: str,
    has_cost_subtab: bool = False,
    topup_pool: pd.DataFrame = None,
):
    """
    Renders a full packaging tab (controls → generate → metrics → tables → download).
    Optionally wraps generation in a subtab and adds a Package Costing subtab.
    """

    def _render_generation():
        if stream_df.empty:
            st.info(empty_msg)
            return

        # ── Controls ────────────────────────────────────────────────────────
        col1, col2, col3, _ = st.columns([1, 1.4, 0.8, 2])

        with col1:
            cost_mode = st.selectbox(
                "Cost mode",
                ["median", "lowest", "contractor"],
                key=f"{state_key}_cost_mode",
            )
        with col2:
            mode_options = ["Max package value", "Pipes per package"]
            if state_key == "rec":
                mode_options.append("One pipe per package")
            packaging_mode = st.selectbox(
                "Packaging mode",
                mode_options,
                key=f"{state_key}_packaging_mode",
            )
        with col3:
            debug = st.checkbox("Show debug", key=f"{state_key}_debug")



        # ── Size grouping multiselects ──────────────────────────────────────
        if ASSET_DIAM_COL in stream_df.columns:
            available_diams_grp = sorted(
                stream_df[ASSET_DIAM_COL]
                .dropna()
                .apply(lambda x: int(float(x)))
                .unique()
                .tolist()
            )
            diam_options = [str(d) for d in available_diams_grp]
            st.markdown("**Size groups** — diameters selected in the same group can share packages. Group 1 takes priority over Group 2, etc.")
            grp_col1, grp_col2 = st.columns(2)
            with grp_col1:
                grp1 = st.multiselect("Size group 1", options=diam_options, default=[], key=f"{state_key}_diam_grp1")
                grp3 = st.multiselect("Size group 3", options=diam_options, default=[], key=f"{state_key}_diam_grp3")
            with grp_col2:
                grp2 = st.multiselect("Size group 2", options=diam_options, default=[], key=f"{state_key}_diam_grp2")
                grp4 = st.multiselect("Size group 4", options=diam_options, default=[], key=f"{state_key}_diam_grp4")

            # Build ordered list of groups, first group takes priority for any overlap
            raw_groups = [[int(d) for d in g] for g in [grp1, grp2, grp3, grp4] if g]
            # Resolve overlaps — first group wins
            seen = set()
            custom_groups = []
            for grp in raw_groups:
                resolved = [d for d in grp if d not in seen]
                seen.update(resolved)
                if resolved:
                    custom_groups.append(resolved)
        else:
            custom_groups = []

        # ── Proximity grouping ──────────────────────────────────────────────
        has_coords = X_MID_COL in stream_df.columns and Y_MID_COL in stream_df.columns
        if has_coords:
            proximity_options = ["Suburb (no coordinates)"] + [f"{p}m" for p in PROXIMITY_OPTIONS]
        else:
            proximity_options = ["Suburb (no coordinates)"]
        proximity = st.selectbox(
            "Grouping method",
            proximity_options,
            key=f"{state_key}_proximity",
            help="Choose how pipes are grouped into packages. Coordinate-based grouping clusters nearby pipes regardless of suburb.",
        )

        contractor = None
        if cost_mode == "contractor":
            contractors = get_contractors(rates)
            contractor = st.selectbox(
                "Contractor",
                contractors if contractors else ["No contractors found"],
                key=f"{state_key}_contractor",
            )

        if packaging_mode == "Max package value":
            max_value = st.number_input(
                "Max package value ($)",
                min_value=1000.0,
                value=50000.0,
                step=1000.0,
                key=f"{state_key}_max_value",
            )
            pipes_per_package = None
        elif packaging_mode == "One pipe per package":
            pipes_per_package = 1
            max_value = None
        else:
            pipes_per_package = st.number_input(
                "Pipes per package",
                min_value=1,
                value=10,
                step=1,
                key=f"{state_key}_pipes_per_package",
            )
            max_value = None

        # ── Cost adjustments ────────────────────────────────────────────────
        st.markdown("**Cost Adjustments**")
        cadj1, cadj2 = st.columns(2)
        with cadj1:
            traffic_control = st.checkbox(
                "Add provisional traffic control ($2,000/pipe)",
                value=True,
                key=f"{state_key}_traffic_control",
            )
        with cadj2:
            use_initiation = st.checkbox(
                "Apply project initiation multiplier (×1.15)",
                value=True,
                key=f"{state_key}_initiation",
            )

        topup_mode = "No top up"
        if packaging_mode == "Max package value" and state_key != "rec":
            topup_mode = st.radio(
                "Top up underfilled packages",
                ["No top up", "Same condition only", "Same condition first, lower condition if needed"],
                key=f"{state_key}_topup_mode",
                horizontal=True,
            )

        # ── Generate ────────────────────────────────────────────────────────
        if st.button(f"Generate {tab_label.lower()} packages", key=f"{state_key}_generate"):
            df = stream_df.copy()
            df[ASSET_ID_COL] = df[ASSET_ID_COL].astype(str)



            TRAFFIC_CONTROL_PER_PIPE = 2000.0
            INITIATION_MULTIPLIER = 1.15

            if state_key == "rec":
                # Full reconstruction cost breakdown (pipe + excavation + backfill + demolition + waste)
                cost_results = df.apply(
                    lambda r: calc_reconstruction_cost(
                        diameter_mm=r.get(ASSET_DIAM_COL),
                        length_m=r.get(ASSET_LEN_COL),
                        rates=rates,
                        mode="vendor" if cost_mode == "contractor" else cost_mode,
                        vendor=contractor if cost_mode == "contractor" else None,
                        pipe_class=DEFAULT_PIPE_CLASS,
                    ),
                    axis=1,
                )
                df["rate_per_m"] = cost_results.apply(lambda x: x["pipe_rate"] if x else None)
                df["pipe_cost"] = cost_results.apply(lambda x: x["total_cost"] if x else None)
                df["_exc_cost"] = cost_results.apply(lambda x: x["excavation_cost"] if x else None)
                df["_bkf_cost"] = cost_results.apply(lambda x: x["backfill_cost"] if x else None)
                df["_dem_cost"] = cost_results.apply(lambda x: x["demolition_cost"] if x else None)
                df["_waste_cost"] = cost_results.apply(lambda x: x["waste_cost"] if x else None)
                df["_trench_vol"] = cost_results.apply(lambda x: x["trench_volume_m3"] if x else None)
                df["_waste_t"] = cost_results.apply(lambda x: x["waste_weight_t"] if x else None)
            else:
                df["rate_per_m"] = df.apply(
                    lambda r: rate_fn(
                        r.get(ASSET_DIAM_COL),
                        r.get(ASSET_LEN_COL),
                        rates,
                        mode="vendor" if cost_mode == "contractor" else cost_mode,
                        vendor=contractor if cost_mode == "contractor" else None,
                    ),
                    axis=1,
                )
                df["pipe_cost"] = df["rate_per_m"] * df[ASSET_LEN_COL]

            # Store uncosted pipes in session state so they display after button press
            uncosted_df = df[df["pipe_cost"].isna()].copy()
            st.session_state[f"{state_key}_uncosted"] = uncosted_df

            # For the relining stream: condition 8 pipes with no matching rate
            # (e.g. too small to reline) are redirected to reconstruction.
            if state_key == "rel" and not uncosted_df.empty:
                existing_rec = st.session_state.get("rec_overflow", pd.DataFrame())
                st.session_state["rec_overflow"] = pd.concat(
                    [existing_rec, uncosted_df], ignore_index=True
                ).drop_duplicates(subset=[ASSET_ID_COL])
                st.info(
                    f"{len(uncosted_df)} pipe(s) with no relining rate have been "
                    f"added to the Reconstruction tab for costing."
                )

            # Remove uncosted pipes from the df used for packaging
            df = df[df["pipe_cost"].notna()].copy()

            if state_key != "rec":
                if traffic_control:
                    df["pipe_cost"] = df["pipe_cost"] + TRAFFIC_CONTROL_PER_PIPE
                if use_initiation:
                    df["pipe_cost"] = df["pipe_cost"] * INITIATION_MULTIPLIER
            else:
                # For reconstruction, traffic control and initiation apply to total cost
                if traffic_control:
                    df["pipe_cost"] = df["pipe_cost"] + TRAFFIC_CONTROL_PER_PIPE
                if use_initiation:
                    df["pipe_cost"] = df["pipe_cost"] * INITIATION_MULTIPLIER

            suburb_col = resolve_suburb_column(df)
            st.session_state[f"{state_key}_suburb_col"] = suburb_col

            # Only mark as TBA if pipe has no suburb AND no coordinates
            # Pipes with coordinates but no suburb can still be spatially grouped
            TBA_VALUES = {"tba", "to be determined", "n/a", "not applicable", "", "nan", "none"}
            if suburb_col and suburb_col in df.columns:
                df[suburb_col] = df[suburb_col].astype(str).str.strip()
                suburb_is_tba = df[suburb_col].str.lower().isin(TBA_VALUES)
                has_coords = (
                    df[X_MID_COL].notna() & df[Y_MID_COL].notna()
                    if X_MID_COL in df.columns and Y_MID_COL in df.columns
                    else pd.Series(False, index=df.index)
                )
                # Only force TBA label if no suburb AND no coordinates
                truly_tba = suburb_is_tba & ~has_coords
                df.loc[truly_tba, suburb_col] = "TBA"

            # Apply spatial clustering if a proximity option is selected
            use_proximity = proximity != "Suburb (no coordinates)"
            if use_proximity:
                radius_m = int(''.join(filter(str.isdigit, proximity)))
                df = assign_spatial_clusters(
                    df,
                    radius_m=radius_m,
                    suburb_col=suburb_col if suburb_col else SUBURB_COL,
                )
                group_col = "_cluster"
            else:
                group_col = None

            if packaging_mode == "Max package value":
                packaged = split_into_value_packages_twop(
                    df,
                    max_package_value=max_value,
                    suburb_col=suburb_col,
                    prefix=prefix,
                    cost_col="pipe_cost",
                    adjacent_fill=(topup_mode != "No top up"),
                    custom_groups=custom_groups if custom_groups else None,
                    group_col=group_col,
                )
                # Top up underfilled packages based on selected mode
                if topup_mode != "No top up":

                    def build_costed_pool(source_df):
                        """Cost a pool of pipes using the same settings as the main stream."""
                        p = source_df.copy()
                        p[ASSET_ID_COL] = p[ASSET_ID_COL].astype(str)
                        p["rate_per_m"] = p.apply(
                            lambda r: rate_fn(
                                r.get(ASSET_DIAM_COL),
                                r.get(ASSET_LEN_COL),
                                rates,
                                mode="vendor" if cost_mode == "contractor" else cost_mode,
                                vendor=contractor if cost_mode == "contractor" else None,
                            ),
                            axis=1,
                        )
                        p = p[p["rate_per_m"].notna()].copy()
                        p["pipe_cost"] = p["rate_per_m"] * p[ASSET_LEN_COL]
                        if traffic_control:
                            p["pipe_cost"] = p["pipe_cost"] + TRAFFIC_CONTROL_PER_PIPE
                        if use_initiation:
                            p["pipe_cost"] = p["pipe_cost"] * INITIATION_MULTIPLIER
                        if ASSET_DIAM_COL in p.columns:
                            p["_size_group"] = p[ASSET_DIAM_COL].apply(
                                lambda d: get_package_group_label(float(d)) if pd.notna(d) else "unknown"
                            )
                        # Exclude pipes already in the main stream
                        already = set(df[ASSET_ID_COL].astype(str))
                        return p[~p[ASSET_ID_COL].isin(already)]

                    # Same condition pool = unpackaged pipes from the same stream_df
                    packaged_ids = set(packaged[ASSET_ID_COL].astype(str))
                    same_cond_unpackaged = stream_df[
                        ~stream_df[ASSET_ID_COL].astype(str).isin(packaged_ids)
                    ]
                    same_cond_pool = build_costed_pool(same_cond_unpackaged) if not same_cond_unpackaged.empty else pd.DataFrame()

                    if topup_mode == "Same condition only":
                        if not same_cond_pool.empty:
                            packaged = topup_packages(
                                packaged=packaged,
                                topup_pool=same_cond_pool,
                                max_package_value=max_value,
                                suburb_col=suburb_col,
                            )

                    elif topup_mode == "Same condition first, lower condition if needed":
                        # First pass — same condition
                        if not same_cond_pool.empty:
                            packaged = topup_packages(
                                packaged=packaged,
                                topup_pool=same_cond_pool,
                                max_package_value=max_value,
                                suburb_col=suburb_col,
                            )
                        # Second pass — lower condition for still-underfilled packages
                        if topup_pool is not None and not topup_pool.empty:
                            lower_pool = build_costed_pool(topup_pool)
                            if not lower_pool.empty:
                                packaged = topup_packages(
                                    packaged=packaged,
                                    topup_pool=lower_pool,
                                    max_package_value=max_value,
                                    suburb_col=suburb_col,
                                )
            else:
                packaged = split_into_packages_by_count(
                    df,
                    pipes_per_package=int(pipes_per_package),
                    group_cols=group_cols,
                    prefix=prefix,
                )

            # Store pre-topup version for separate download if topup was applied
            st.session_state[f"{state_key}_packaged_base"] = packaged.copy()

            summary = summarise_pipe_packages(packaged)

            packaged[ASSET_ID_COL] = (
                packaged[ASSET_ID_COL]
                .astype(str)
                .str.replace(r"\.0$", "", regex=True)
            )

            st.session_state[f"{state_key}_packaged"] = packaged
            st.session_state[f"{state_key}_summary"] = summary

            st.success(f"{packaged['package_id'].nunique()} {tab_label.lower()} packages created.")

        # ── Uncosted pipes warning ───────────────────────────────────────────
        uncosted_df = st.session_state.get(f"{state_key}_uncosted")
        if uncosted_df is not None and not uncosted_df.empty and state_key != "rel":
            st.warning(f"{len(uncosted_df)} pipe(s) had no matching rate and have been excluded from cost totals.")
            _suburb_col = st.session_state.get(f"{state_key}_suburb_col")
            uncosted_df[ASSET_ID_COL] = uncosted_df[ASSET_ID_COL].astype(str).str.replace(r"\.0$", "", regex=True)
            uncosted_cols = [c for c in [ASSET_LEN_COL, _suburb_col, ASSET_DIAM_COL, ASSET_ID_COL] if c and c in uncosted_df.columns]
            st.dataframe(uncosted_df[uncosted_cols].reset_index(drop=True).rename(lambda x: x+1), use_container_width=True)

        # ── Display results ─────────────────────────────────────────────────
        packaged = st.session_state.get(f"{state_key}_packaged")
        suburb_col = st.session_state.get(f"{state_key}_suburb_col")

        if packaged is not None and not packaged.empty:
            total_pipes = len(packaged)
            total_len = packaged[ASSET_LEN_COL].sum()
            total_cost = packaged["pipe_cost"].sum()
            num_packages = packaged["package_id"].nunique()

            k1, k2, k3, k4 = st.columns(4)
            k1.metric("Packages", num_packages)
            k2.metric("Pipes", total_pipes)
            k3.metric("Length (m)", f"{total_len:,.1f}")
            k4.metric("Cost ($)", f"{total_cost:,.0f}")

            st.subheader(f"Pipe {tab_label} Details")

            display_cols = [ASSET_LEN_COL, "pipe_cost", ASSET_DIAM_COL, ASSET_ID_COL, "package_id"]
            if suburb_col and suburb_col in packaged.columns:
                display_cols.insert(2, suburb_col)
            existing_display = [c for c in display_cols if c in packaged.columns]
            st.dataframe(packaged[existing_display].reset_index(drop=True).rename(lambda x: x+1), use_container_width=True)

            # ── Package summary table ────────────────────────────────────────
            st.subheader(f"{tab_label} Package Summary")

            pkg_summary_rows = []
            for pkg_id, group in packaged.groupby("package_id"):
                asset_ids = (
                    group[ASSET_ID_COL]
                    .dropna()
                    .astype(str)
                    .str.replace(r"\.0$", "", regex=True)
                    .tolist()
                )

                try:
                    diameters = sorted({str(int(float(d))) for d in group[ASSET_DIAM_COL].unique()})
                except (ValueError, TypeError):
                    diameters = sorted(group[ASSET_DIAM_COL].astype(str).unique())

                row = {"package_id": pkg_id}
                if suburb_col and suburb_col in group.columns:
                    row[suburb_col] = ", ".join(sorted(group[suburb_col].astype(str).unique()))
                row[ASSET_DIAM_COL] = ", ".join(diameters)
                row[ASSET_LEN_COL] = group[ASSET_LEN_COL].sum()
                row["pipe_cost"] = group["pipe_cost"].sum()
                row[ASSET_ID_COL] = ", ".join(asset_ids)
                pkg_summary_rows.append(row)

            pkg_summary_df = pd.DataFrame(pkg_summary_rows)
            if "pipe_cost" in pkg_summary_df.columns:
                pkg_summary_df = pkg_summary_df.rename(columns={"pipe_cost": "Total Cost ($)"})
                pkg_summary_df["Total Cost ($)"] = pkg_summary_df["Total Cost ($)"].apply(lambda x: f"${x:,.0f}")
            if ASSET_LEN_COL in pkg_summary_df.columns:
                pkg_summary_df[ASSET_LEN_COL] = pkg_summary_df[ASSET_LEN_COL].apply(lambda x: f"{x:,.1f}")
            st.dataframe(pkg_summary_df.reset_index(drop=True).rename(lambda x: x+1), use_container_width=True)

            st.subheader(f"Download {tab_label.lower()} packages")
            zip_buffer = create_zip_from_packages(
                packaged,
                suburb_col=suburb_col,
                diam_col=ASSET_DIAM_COL if ASSET_DIAM_COL in packaged.columns else None,
            )
            st.download_button(
                f"Download all {tab_label.lower()} packages (ZIP)",
                data=zip_buffer,
                file_name=zip_filename,
                mime="application/zip",
                key=f"{state_key}_download_zip",
            )

            # Offer a separate download for base (no top-up) packages if top-up was applied
            packaged_base = st.session_state.get(f"{state_key}_packaged_base")
            if (
                packaged_base is not None
                and not packaged_base.empty
                and topup_mode != "No top up"
                and len(packaged_base) != len(packaged)
            ):
                base_zip_buffer = create_zip_from_packages(
                    packaged_base,
                    suburb_col=suburb_col,
                    diam_col=ASSET_DIAM_COL if ASSET_DIAM_COL in packaged_base.columns else None,
                )
                base_zip_name = zip_filename.replace(".zip", "_base_no_topup.zip")
                st.download_button(
                    f"Download base packages only (no top-up) (ZIP)",
                    data=base_zip_buffer,
                    file_name=base_zip_name,
                    mime="application/zip",
                    key=f"{state_key}_download_zip_base",
                )

            if debug:
                st.subheader("Debug data")
                st.dataframe(packaged.reset_index(drop=True).rename(lambda x: x+1), use_container_width=True)

    if has_cost_subtab:
        subtab_gen, subtab_cost, subtab_edit = st.tabs(["Package Generation", "Package Costing", "Edit Packages"])
        with subtab_gen:
            _render_generation()
        with subtab_cost:
            render_costing_subtab(
                state_key=state_key,
                rates=rates,
                rate_fn=rate_fn,
                label=tab_label,
            )
        with subtab_edit:
            render_edit_subtab(state_key=state_key, tab_label=tab_label, prefix=prefix)
    else:
        subtab_gen, subtab_edit = st.tabs(["Package Generation", "Edit Packages"])
        with subtab_gen:
            _render_generation()
        with subtab_edit:
            render_edit_subtab(state_key=state_key, tab_label=tab_label, prefix=prefix)



def render_edit_subtab(state_key: str, tab_label: str, prefix: str):
    """Renders the package editing subtab for a given stream."""
    packaged = st.session_state.get(f"{state_key}_packaged")

    if packaged is None or packaged.empty:
        st.info("Generate packages first before editing.")
        return

    # Work on a copy stored in session state so edits persist
    edit_key = f"{state_key}_packaged_edited"
    if edit_key not in st.session_state:
        st.session_state[edit_key] = packaged.copy()

    edited = st.session_state[edit_key]

    # If base packages have been regenerated, reset edits
    if set(packaged["package_id"].unique()) != set(edited["package_id"].unique()):
        st.session_state[edit_key] = packaged.copy()
        edited = st.session_state[edit_key]

    all_packages = sorted(edited["package_id"].unique())

    st.markdown("Select a package to view its pipes and move them to another package.")

    col1, col2 = st.columns([2, 3])
    with col1:
        selected_pkg = st.selectbox(
            "Select package to edit",
            all_packages,
            key=f"{state_key}_edit_select_pkg",
        )

    pkg_df = edited[edited["package_id"] == selected_pkg].copy()

    if pkg_df.empty:
        st.warning("No pipes in this package.")
        return

    # Show package summary
    suburb_col = st.session_state.get(f"{state_key}_suburb_col")
    pipe_count = len(pkg_df)
    total_cost = pkg_df["pipe_cost"].sum() if "pipe_cost" in pkg_df.columns else 0
    total_len = pkg_df[ASSET_LEN_COL].sum() if ASSET_LEN_COL in pkg_df.columns else 0

    m1, m2, m3 = st.columns(3)
    m1.metric("Pipes", pipe_count)
    m2.metric("Length (m)", f"{total_len:,.1f}")
    m3.metric("Cost ($)", f"${total_cost:,.0f}")

    st.markdown("**Pipes in this package** — use the dropdown on each row to move a pipe:")

    # Other packages available as move targets (plus "— keep here —")
    other_packages = ["— keep here —"] + [p for p in all_packages if p != selected_pkg] + [f"NEW PACKAGE"]

    # Build per-pipe move dropdowns
    moves = {}  # asset_index -> target_package
    display_cols = [c for c in [ASSET_ID_COL, suburb_col, ASSET_DIAM_COL, ASSET_LEN_COL, "pipe_cost"] if c and c in pkg_df.columns]

    for idx, row in pkg_df.iterrows():
        asset_id = str(row[ASSET_ID_COL]) if ASSET_ID_COL in pkg_df.columns else str(idx)
        cols = st.columns([2, 2, 2, 2, 3])
        for ci, col_name in enumerate(display_cols[:4]):
            val = row[col_name]
            if col_name == "pipe_cost":
                cols[ci].markdown(f"**{col_name}:** ${val:,.0f}")
            elif col_name == ASSET_LEN_COL:
                cols[ci].markdown(f"**{col_name}:** {val:,.1f}m")
            elif col_name == ASSET_DIAM_COL:
                try:
                    cols[ci].markdown(f"**{col_name}:** {int(float(val))}mm")
                except (ValueError, TypeError):
                    cols[ci].markdown(f"**{col_name}:** {val}")
            else:
                cols[ci].markdown(f"**{col_name}:** {val}")
        move_to = cols[4].selectbox(
            "Move to",
            other_packages,
            key=f"{state_key}_move_{idx}",
            label_visibility="collapsed",
        )
        moves[idx] = move_to

    st.markdown("---")
    if st.button("Apply moves", key=f"{state_key}_apply_moves"):
        new_pkg_counter = None  # Lazy-create new package ID if needed

        for idx, target in moves.items():
            if target == "— keep here —":
                continue
            if target == "NEW PACKAGE":
                if new_pkg_counter is None:
                    # Find next available package number
                    existing_nums = []
                    for pid in edited["package_id"].unique():
                        try:
                            existing_nums.append(int(pid.split("_")[-1]))
                        except ValueError:
                            pass
                    new_pkg_counter = max(existing_nums, default=0) + 1
                    new_pkg_id = f"{prefix}{new_pkg_counter:03d}"
                else:
                    new_pkg_id = f"{prefix}{new_pkg_counter:03d}"
                edited.loc[idx, "package_id"] = new_pkg_id
            else:
                edited.loc[idx, "package_id"] = target

        # Remove any packages that are now empty (all pipes moved out)
        # then renumber all packages in suburb-alphabetical order
        if suburb_col and suburb_col in edited.columns:
            edited = edited.sort_values([suburb_col, "package_id"])
        else:
            edited = edited.sort_values("package_id")

        # Renumber: assign sequential IDs in current sort order
        old_ids = list(dict.fromkeys(edited["package_id"]))  # preserve order, deduplicate
        renumber_map = {old_id: f"{prefix}{i+1:03d}" for i, old_id in enumerate(old_ids)}
        edited["package_id"] = edited["package_id"].map(renumber_map)

        st.session_state[edit_key] = edited
        st.success("Moves applied and packages renumbered.")
        st.rerun()

    # Download edited packages
    st.markdown("---")
    st.subheader("Download edited packages")
    suburb_col_val = st.session_state.get(f"{state_key}_suburb_col")
    zip_buffer = create_zip_from_packages(
        edited,
        suburb_col=suburb_col_val,
        diam_col=ASSET_DIAM_COL if ASSET_DIAM_COL in edited.columns else None,
    )
    st.download_button(
        f"Download edited {tab_label.lower()} packages (ZIP)",
        data=zip_buffer,
        file_name=f"{prefix.lower().rstrip('_')}_packages_edited.zip",
        mime="application/zip",
        key=f"{state_key}_download_edited_zip",
    )

    # Reset button
    if st.button("Reset all edits", key=f"{state_key}_reset_edits"):
        if edit_key in st.session_state:
            del st.session_state[edit_key]
        st.rerun()

# ---------------------------------------------------------
# LOAD DATA
# ---------------------------------------------------------

@st.cache_data
def load_assets():
    return pd.read_csv(ASSETS_PATH, low_memory=False)


@st.cache_data
def load_rates():
    return load_relining_rates(), load_reconstruction_rates()


assets_df = load_assets()
relining_rates, reconstruction_rates = load_rates()
rel_df, rec_df, amp_df = split_streams(assets_df)

# Debug — remove once confirmed working
if "Pipe_Start_Address" not in assets_df.columns:
    st.warning(f"Pipe_Start_Address column not found in: {ASSETS_PATH}")

# ---------------------------------------------------------
# SESSION STATE
# ---------------------------------------------------------

def init_state():
    keys = [
        "rel_packaged", "rel_summary", "rel_suburb_col",
        "rec_packaged", "rec_summary", "rec_suburb_col",
        "amp_packaged", "amp_summary", "amp_suburb_col",
        "rel_uncosted", "rec_uncosted", "amp_uncosted",
    ]
    for k in keys:
        if k not in st.session_state:
            st.session_state[k] = None

init_state()

# ---------------------------------------------------------
# MAIN UI
# ---------------------------------------------------------

st.title("Stormwater Packaging Tool")

tab1, tab2, tab3 = st.tabs([
    "Relining (Cond 7–8)",
    "Reconstruction (Cond 9–10)",
    "Amplification",
])

with tab1:
    st.header("Relining")
    rel_cond_option = st.radio(
        "Include conditions",
        ["Condition 8 only", "Condition 7 only", "Condition 7 and 8"],
        horizontal=True,
        key="rel_condition_filter",
    )
    # Re-filter rel_df based on selection
    raw_cond = pd.to_numeric(
        assets_df[CONDITION_COL].astype(str).str.extract(r"(\d+)")[0],
        errors="coerce"
    ) if CONDITION_COL in assets_df.columns else pd.Series(dtype=float)
    if rel_cond_option == "Condition 8 only":
        rel_df_filtered = assets_df[raw_cond == 8].copy()
        rel_empty_msg = "No condition 8 pipes found."
    elif rel_cond_option == "Condition 7 only":
        rel_df_filtered = assets_df[raw_cond == 7].copy()
        rel_empty_msg = "No condition 7 pipes found."
    else:
        rel_df_filtered = assets_df[raw_cond.isin([7, 8])].copy()
        rel_empty_msg = "No condition 7 or 8 pipes found."

    # Build top-up pool: cond 7 pipes when running cond 8, and vice versa
    if rel_cond_option == "Condition 8 only":
        rel_topup_pool = assets_df[raw_cond == 7].copy()
    elif rel_cond_option == "Condition 7 only":
        rel_topup_pool = assets_df[raw_cond == 8].copy()
    else:
        rel_topup_pool = pd.DataFrame()  # Both already included — nothing to top up

    render_packaging_tab(
        stream_df=rel_df_filtered,
        rates=relining_rates,
        rate_fn=get_relining_rate,
        prefix="RLN_",
        state_key="rel",
        empty_msg=rel_empty_msg,
        zip_filename="relining_packages.zip",
        tab_label="Relining",
        has_cost_subtab=True,
        topup_pool=rel_topup_pool,
    )

with tab2:
    st.header("Reconstruction Packages")
    # Merge any condition 8 pipes that couldn't be relined into this stream
    rec_overflow = st.session_state.get("rec_overflow", pd.DataFrame())
    rec_df_combined = pd.concat([rec_df, rec_overflow], ignore_index=True).drop_duplicates(subset=[ASSET_ID_COL]) if not rec_overflow.empty else rec_df
    render_packaging_tab(
        stream_df=rec_df_combined,
        rates=reconstruction_rates,
        rate_fn=get_reconstruction_rate,
        prefix="REC_",
        state_key="rec",
        empty_msg="No condition 9–10 pipes found.",
        zip_filename="reconstruction_packages.zip",
        tab_label="Reconstruction",
        topup_pool=assets_df[raw_cond.isin([7, 8])].copy() if CONDITION_COL in assets_df.columns else pd.DataFrame(),
    )

with tab3:
    st.header("Amplification Packages")
    render_packaging_tab(
        stream_df=amp_df,
        rates=reconstruction_rates,  # Amplification uses reconstruction rates
        rate_fn=get_reconstruction_rate,
        prefix="AMP_",
        state_key="amp",
        empty_msg="No amplification candidates found.",
        zip_filename="amplification_packages.zip",
        tab_label="Amplification",
    )