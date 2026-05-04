"""
populate_forward_works.py

Creates a standalone Excel sheet ready to copy into the
15) Pipe Renewal Backlog tab of the forward works program.

Packages are assigned to years in order, filling each year's budget
as close as possible to the target. Budget escalates from $3M in Year 1
to cover the full cost of all packages by end of 2031-2032.

Usage:
    py populate_forward_works.py

Files needed in packaging/data/:
    relining_packages.zip   — zip of package CSVs from the packaging tool

Output:
    packaging/data/pipe_renewal_backlog.xlsx
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import zipfile, re
from pathlib import Path

PACKAGING_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = PACKAGING_ROOT / "data"

ZIP_FILE    = str(DATA_DIR / "relining_packages.zip")
OUTPUT_FILE = str(DATA_DIR / "pipe_renewal_backlog.xlsx")

YEARS = ["2027-2028", "2028-2029", "2029-2030", "2030-2031", "2031-2032"]
BASE_BUDGET = 3_000_000

# ---------------------------------------------------------
# LOAD PACKAGES
# ---------------------------------------------------------
print("Loading package data...")
all_pipes = []
with zipfile.ZipFile(ZIP_FILE, 'r') as z:
    for name in z.namelist():
        if name.endswith('.csv') and re.match(r'.*RLN_\d+.*\.csv', name):
            with z.open(name) as f:
                df = pd.read_csv(f, dtype=str)
                df = df[df.get('package_id', pd.Series()).ne('Totals')]
                if not df.empty:
                    all_pipes.append(df)

all_pipes = pd.concat(all_pipes, ignore_index=True)

def clean_list(vals, exclude=('', 'nan', 'tba', 'none')):
    return sorted(set(str(v).strip() for v in vals
                      if pd.notna(v) and str(v).strip().lower() not in exclude))

summary = all_pipes.groupby('package_id').apply(lambda g: pd.Series({
    'suburbs': ', '.join(clean_list(g.get('Asset Suburb', []))),
    'streets': ', '.join(clean_list(
        str(v).split(',')[0].strip() for v in g.get('Pipe_Start_Address', [])
        if pd.notna(v)
    )),
    'assets':  '\n'.join(str(v).replace('.0','') for v in g.get('Asset', []) if pd.notna(v)),
    'cost':    pd.to_numeric(
        g.get('pipe_cost', pd.Series()).astype(str).str.replace(r'[$,]','',regex=True),
        errors='coerce'
    ).sum(),
})).reset_index()

summary = summary[summary['suburbs'].str.strip().ne('')]
summary = summary[~summary['package_id'].str.upper().str.contains('TBA')]
summary = summary.sort_values('package_id').reset_index(drop=True)

total_cost = summary['cost'].sum()
n = len(YEARS)
print(f"  {len(summary)} packages, total ${total_cost:,.0f}")

# ---------------------------------------------------------
# CALCULATE ESCALATING ANNUAL BUDGETS
# ---------------------------------------------------------
# Arithmetic sequence starting at BASE_BUDGET, summing to total_cost
# sum = n/2 * (2a + (n-1)d) => d = (total - n*a) / (n*(n-1)/2)
d = (total_cost - BASE_BUDGET * n) / (n * (n - 1) / 2)
budgets = {y: BASE_BUDGET + i * d for i, y in enumerate(YEARS)}

print("  Annual budgets:")
for y, b in budgets.items():
    print(f"    {y}: ${b:,.0f}")

# ---------------------------------------------------------
# ASSIGN PACKAGES TO YEARS (sequential fill)
# ---------------------------------------------------------
year_idx = 0
year_running = 0.0
assignments = []
year_totals = {y: 0.0 for y in YEARS}

print(f"  First 5 package costs: {summary['cost'].head().tolist()}")
print(f"  Year 1 budget: ${budgets[YEARS[0]]:,.0f}")

for _, pkg in summary.iterrows():
    # Move to next year if adding this package would exceed current year budget by >5%
    if year_idx < len(YEARS) - 1 and year_running + pkg['cost'] > budgets[YEARS[year_idx]] * 1.05:
        year_idx += 1
        year_running = 0.0
    year = YEARS[year_idx]
    assignments.append(year)
    year_running += pkg['cost']
    year_totals[year] += pkg['cost']

summary['year'] = assignments

print("  Final distribution:")
for y in YEARS:
    pkgs = summary[summary['year'] == y]
    print(f"    {y}: {len(pkgs)} packages, ${year_totals[y]:,.0f} (budget ${budgets[y]:,.0f})")

# ---------------------------------------------------------
# BUILD EXCEL
# ---------------------------------------------------------
GREEN = "0A7A3B"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "15) Pipe Renewal Backlog"

headers = [
    "Package", "Program Description", "Brief Required?", "Budget ($)",
    "Short description of work", "Suburb(s)", "Street(s)", "Ward",
    "Stage by end of FY26/27", "Asset number",
    "2027-28 Q1", "2027-28 Q2", "2027-28 Q3", "2027-28 Q4",
    "2028-29 Q1", "2028-29 Q2", "2028-29 Q3", "2028-29 Q4",
    "2029-30 Q1", "2029-30 Q2", "2029-30 Q3", "2029-30 Q4",
    "2030-31 Q1", "2030-31 Q2", "2030-31 Q3", "2030-31 Q4",
    "2031-32 Q1", "2031-32 Q2", "2031-32 Q3", "2031-32 Q4",
]

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=GREEN)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 40

# Year -> Q3 column (1-based)
year_q3 = {
    "2027-2028": 13, "2028-2029": 17,
    "2029-2030": 21, "2030-2031": 25, "2031-2032": 29,
}

# Add a year header row above data
current_year = None
data_row = 2

for _, pkg in summary.iterrows():
    # Insert year group header when year changes
    if pkg['year'] != current_year:
        current_year = pkg['year']
        cell = ws.cell(row=data_row, column=1, value=f"--- {current_year} ---")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="555555")
        budget_cell = ws.cell(row=data_row, column=4,
                              value=f"Budget: ${budgets[current_year]:,.0f}")
        budget_cell.font = Font(bold=True, color="FFFFFF")
        budget_cell.fill = PatternFill("solid", fgColor="555555")
        ws.row_dimensions[data_row].height = 18
        data_row += 1

    ws.cell(row=data_row, column=1,  value=pkg['package_id'])
    ws.cell(row=data_row, column=2,  value='Stormwater Pipe Relining')
    ws.cell(row=data_row, column=3,  value='YES')
    c = ws.cell(row=data_row, column=4, value=round(pkg['cost'], 2))
    c.number_format = '$#,##0'
    ws.cell(row=data_row, column=5,  value='Relining works for poor condition pipes')
    ws.cell(row=data_row, column=6,  value=pkg['suburbs'] or 'LGA Wide')
    ws.cell(row=data_row, column=7,  value=pkg['streets'] or '')
    ws.cell(row=data_row, column=8,  value='LGA Wide')
    ws.cell(row=data_row, column=9,  value='')
    a = ws.cell(row=data_row, column=10, value=pkg['assets'])
    a.alignment = Alignment(wrap_text=True, vertical='top')

    q3 = year_q3.get(pkg['year'])
    if q3:
        ws.cell(row=data_row, column=q3,     value=3)
        ws.cell(row=data_row, column=q3 + 1, value=4)

    ws.row_dimensions[data_row].height = max(15, len(pkg['assets'].split('\n')) * 12)
    data_row += 1

# Column widths
for col, width in zip(range(1, 11), [15, 25, 12, 14, 35, 35, 45, 12, 20, 35]):
    ws.column_dimensions[ws.cell(1, col).column_letter].width = width
for col in range(11, 31):
    ws.column_dimensions[ws.cell(1, col).column_letter].width = 7

wb.save(OUTPUT_FILE)
print(f"\nSaved to: {OUTPUT_FILE}")