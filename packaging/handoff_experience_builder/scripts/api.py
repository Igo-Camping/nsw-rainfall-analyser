from __future__ import annotations

import io
import os
import zipfile
from functools import lru_cache
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel

from cost_engine import (
    ASSET_DIAM_COL,
    ASSET_ID_COL,
    ASSET_LEN_COL,
    CONDITION_COL,
    MIN_RELINING_DIAMETER_MM,
    X_MID_COL,
    Y_MID_COL,
    add_package_priority,
    assign_spatial_clusters,
    coerce_diameter_value,
    ensure_suburb_column,
    extract_condition_scores,
    package_diameters_text,
    package_max_distance_m,
    PROXIMITY_OPTIONS,
    get_relining_rate,
    get_package_group_label,
    load_relining_rates,
    normalize_asset_columns,
    reindex_packages_by_priority,
    resolve_suburb_column,
    split_into_packages_by_count,
    split_into_value_packages_twop,
    split_streams,
    sort_packages_by_priority,
    summarise_pipe_packages,
    topup_packages,
)


SHARED_ROOT_ENV = "PACKAGING_SHARED_ROOT"
DEFAULT_SHARED_SEGMENTS = [
    "OneDrive - Northern Beaches Council",
    "Stormwater Engineering - General",
    "Pipe Relining",
]

OUTPUT_KEEP_COLS = [
    "package_id",
    "package_priority_rank",
    "package_priority_score",
    "package_total_pipe_priority",
    "package_high_crit_count",
    "package_large_diam_count",
    "package_worst_condition",
    "package_total_length",
    "package_pipe_count",
    "Asset",
    "SWP_Pipe Diameter_mm",
    "Spatial Length_m",
    "SW_Condition",
    "Score",
    "criticality_1dp",
    "diameter_weight",
    "condition_score",
    "pipe_priority_score",
    "SW LGA 20% H1-H6",
    "SWP_ Pipe Material",
    "SWP_ Pipe Class",
    "SW_Observation Metho",
    "SW_Upstream Node",
    "SW_Downstream Node",
    "rate_per_m",
    "pipe_cost",
    "Pipe_Start_Address",
    "Asset Suburb",
    "Diameters (mm)",
    "Max Distance Between Pipes (km)",
    "Number of Pipes",
]


def _shared_root() -> Path:
    configured = os.getenv(SHARED_ROOT_ENV, "").strip()
    if configured:
        return Path(configured).expanduser()

    default_root = Path.home().joinpath(*DEFAULT_SHARED_SEGMENTS)
    if default_root.is_dir():
        return default_root

    return Path(__file__).resolve().parents[1]


def _data_dir() -> Path:
    candidates = [
        _shared_root() / "Data",
        _shared_root() / "data",
        Path(__file__).resolve().parents[1] / "Data",
        Path(__file__).resolve().parents[1] / "data",
    ]
    for candidate in candidates:
        if candidate.is_dir():
            return candidate
    return candidates[0]


def _outputs_dir() -> Path:
    candidates = [
        _shared_root() / "Outputs",
        _shared_root() / "outputs",
        Path(__file__).resolve().parents[1] / "Outputs",
        Path(__file__).resolve().parents[1] / "outputs",
    ]
    for candidate in candidates:
        if candidate.is_dir():
            return candidate
    return candidates[0]


def _pick_existing(candidates: list[Path]) -> Path:
    for candidate in candidates:
        if candidate.is_file():
            return candidate
    return candidates[0]


def get_assets_path() -> Path:
    configured = os.getenv("PACKAGING_ASSETS_PATH", "").strip()
    if configured:
        return Path(configured)
    return _pick_existing(
        [
            _data_dir() / "assets_with_coords.csv",
            _data_dir() / "assets.csv",
        ]
    )


def get_allowed_origins() -> list[str]:
    configured = os.getenv("PACKAGING_ALLOWED_ORIGINS", "").strip()
    if configured:
        return [origin.strip() for origin in configured.split(",") if origin.strip()]
    return [
        "http://localhost",
        "http://localhost:3000",
        "http://localhost:8000",
        "http://localhost:8080",
        "http://127.0.0.1",
        "http://127.0.0.1:3000",
        "http://127.0.0.1:8000",
        "http://127.0.0.1:8080",
        "https://igo-camping.github.io",
    ]


def get_streamlit_url() -> str:
    configured = os.getenv("PACKAGING_STREAMLIT_URL", "").strip()
    if configured:
        return configured
    state_file = Path(__file__).resolve().parent / "web" / "streamlit_url.txt"
    if state_file.is_file():
        file_value = state_file.read_text(encoding="utf-8").lstrip("\ufeff").strip()
        if file_value:
            return file_value
    return "http://localhost:8501"


@lru_cache(maxsize=1)
def load_assets() -> pd.DataFrame:
    return pd.read_csv(get_assets_path(), low_memory=False)


def _value_or_none(value: Any) -> Any:
    if pd.isna(value):
        return None
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return value
    return value


def _stream_summary(name: str, df: pd.DataFrame, suburb_col: str | None) -> dict[str, Any]:
    summary: dict[str, Any] = {
        "name": name,
        "pipe_count": int(len(df)),
        "total_length_m": 0.0,
        "sample_assets": [],
    }

    if df.empty:
        return summary

    if ASSET_LEN_COL in df.columns:
        lengths = pd.to_numeric(df[ASSET_LEN_COL], errors="coerce")
        summary["total_length_m"] = round(float(lengths.fillna(0).sum()), 1)

    if ASSET_DIAM_COL in df.columns:
        diams = pd.to_numeric(df[ASSET_DIAM_COL], errors="coerce").dropna()
        if not diams.empty:
            summary["diameter_range_mm"] = {
                "min": round(float(diams.min()), 1),
                "max": round(float(diams.max()), 1),
            }

    if suburb_col and suburb_col in df.columns:
        top_suburbs = (
            df[suburb_col]
            .fillna("Unknown")
            .astype(str)
            .value_counts()
            .head(5)
            .items()
        )
        summary["top_suburbs"] = [
            {"name": suburb, "pipe_count": int(count)}
            for suburb, count in top_suburbs
        ]

    sample_cols = [
        c
        for c in [ASSET_ID_COL, suburb_col, ASSET_DIAM_COL, ASSET_LEN_COL, CONDITION_COL]
        if c and c in df.columns
    ]
    for _, row in df[sample_cols].head(8).iterrows():
        summary["sample_assets"].append(
            {col: _value_or_none(row[col]) for col in sample_cols}
        )

    return summary


def _split_assets(relining_mode: str) -> dict[str, pd.DataFrame]:
    assets_df = load_assets()
    suburb_col = resolve_suburb_column(assets_df)
    normalized_assets = normalize_asset_columns(assets_df.copy())

    raw_cond = extract_condition_scores(normalized_assets)
    diam = (
        pd.to_numeric(normalized_assets[ASSET_DIAM_COL].apply(coerce_diameter_value), errors="coerce")
        if ASSET_DIAM_COL in normalized_assets.columns
        else pd.Series(dtype=float, index=normalized_assets.index)
    )

    rel_df, rec_df, amp_df = split_streams(assets_df)

    if relining_mode == "Condition 8 only":
        relining_df = normalized_assets[(raw_cond == 8) & (diam >= MIN_RELINING_DIAMETER_MM)].copy()
    elif relining_mode == "Condition 7 only":
        relining_df = normalized_assets[(raw_cond == 7) & (diam >= MIN_RELINING_DIAMETER_MM)].copy()
    else:
        relining_df = normalized_assets[raw_cond.isin([7, 8]) & (diam >= MIN_RELINING_DIAMETER_MM)].copy()

    return {
        "relining": relining_df,
        "relining_cond_8_default": rel_df,
        "reconstruction": rec_df,
        "amplification": amp_df,
        "_suburb_col": suburb_col,
        "_assets_df": normalized_assets,
    }


@lru_cache(maxsize=1)
def load_relining_rates_cached() -> pd.DataFrame:
    return load_relining_rates()


def _cost_relining_df(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    if df.empty:
        empty = df.copy()
        return empty, empty

    rates = load_relining_rates_cached()
    costed = df.copy()
    costed["rate_per_m"] = costed.apply(
        lambda row: get_relining_rate(
            row.get(ASSET_DIAM_COL),
            row.get(ASSET_LEN_COL),
            rates,
            mode="median",
        ),
        axis=1,
    )
    if ASSET_LEN_COL in costed.columns:
        lengths = pd.to_numeric(costed[ASSET_LEN_COL], errors="coerce")
        rates_num = pd.to_numeric(costed["rate_per_m"], errors="coerce")
        costed["pipe_cost"] = rates_num * lengths
    else:
        costed["pipe_cost"] = None

    uncosted = costed[costed["pipe_cost"].isna()].copy()
    costed = costed[costed["pipe_cost"].notna()].copy()
    return costed, uncosted


def get_contractors(rates: pd.DataFrame) -> list[str]:
    exclude_names = ["mean", "average", "median", "stormwater mean"]
    exclude_items = ["establishment", "modifier", "traffic control"]

    for col in ["Component", "Origin of Average/Mean Price"]:
        if rates is None or rates.empty or col not in rates.columns:
            continue

        filtered = rates.copy()
        if "Unit Rate" in filtered.columns:
            filtered["_rate_num"] = pd.to_numeric(filtered["Unit Rate"], errors="coerce").fillna(0)
            if "Item" in filtered.columns:
                item_mask = ~filtered["Item"].astype(str).str.lower().str.contains(
                    "|".join(exclude_items), na=False
                )
                filtered = filtered[item_mask & (filtered["_rate_num"] > 0)]
            else:
                filtered = filtered[filtered["_rate_num"] > 0]

        all_vals = filtered[col].dropna().astype(str).str.strip()
        return sorted(
            v for v in all_vals.unique()
            if not any(excl in v.lower() for excl in exclude_names)
        )

    return []


def _package_summary_rows(df: pd.DataFrame, suburb_col: str | None) -> list[dict[str, Any]]:
    if df.empty or "package_id" not in df.columns:
        return []

    summary = summarise_pipe_packages(df)
    priority_lookup: dict[str, dict[str, Any]] = {}
    ranked = add_package_priority(df.copy()) if not df.empty else df
    if isinstance(ranked, pd.DataFrame) and not ranked.empty and "package_id" in ranked.columns:
        priority_cols = [c for c in ["package_id", "package_priority_rank", "package_priority_score"] if c in ranked.columns]
        if priority_cols:
            for _, item in ranked[priority_cols].drop_duplicates(subset=["package_id"]).iterrows():
                priority_lookup[str(item["package_id"])] = {
                    "package_priority_rank": item.get("package_priority_rank"),
                    "package_priority_score": item.get("package_priority_score"),
                }

    rows: list[dict[str, Any]] = []
    for _, row in summary.iterrows():
        package_id = str(row["package_id"])
        pkg = df[df["package_id"] == package_id]
        priority_meta = priority_lookup.get(package_id, {})
        entry: dict[str, Any] = {
            "package_id": package_id,
            "package_priority_rank": int(_value_or_none(priority_meta.get("package_priority_rank")) or 0),
            "pipe_count": int(_value_or_none(row.get("pipe_count")) or 0),
            "total_length_m": round(float(_value_or_none(row.get("total_length_m")) or 0), 1),
            "total_cost": round(float(_value_or_none(row.get("total_cost")) or 0), 2),
        }
        if _value_or_none(priority_meta.get("package_priority_score")) is not None:
            entry["package_priority_score"] = round(float(_value_or_none(priority_meta.get("package_priority_score")) or 0), 1)
        if suburb_col and suburb_col in pkg.columns:
            suburbs = sorted(
                {
                    value.strip()
                    for value in pkg[suburb_col].fillna("Unknown").astype(str).tolist()
                    if value and value.strip()
                }
            )
            entry["suburb"] = ", ".join(suburbs)
        if ASSET_DIAM_COL in pkg.columns:
            diams = pd.to_numeric(pkg[ASSET_DIAM_COL], errors="coerce").dropna()
            if not diams.empty:
                entry["diameters_mm"] = sorted({int(d) for d in diams})
        rows.append(entry)
    return rows


def format_distance_km(distance_m: float | None) -> str:
    if distance_m is None or pd.isna(distance_m):
        return ""
    return f"{distance_m / 1000:,.1f} km"


def get_package_order(df: pd.DataFrame) -> list[str]:
    if df is None or df.empty or "package_id" not in df.columns:
        return []
    ordered = sort_packages_by_priority(df)
    return list(pd.Index(ordered["package_id"]).drop_duplicates())


def create_zip_from_packages(
    df: pd.DataFrame,
    suburb_col: str | None = None,
    diam_col: str | None = None,
) -> io.BytesIO:
    effective_suburb_col = "Asset Suburb" if "Asset Suburb" in df.columns else suburb_col
    buffer = io.BytesIO()
    csv_filenames: list[str] = []

    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as z:
        for pkg_id in get_package_order(df):
            pkg_df = df[df["package_id"] == pkg_id]
            if pkg_df.empty:
                continue

            parts = [str(pkg_id)]
            if effective_suburb_col and effective_suburb_col in pkg_df.columns:
                suburbs = sorted(
                    {
                        value.strip()
                        for value in pkg_df[effective_suburb_col].fillna("Unknown").astype(str).tolist()
                        if value and value.strip()
                    }
                )
                if suburbs:
                    suburb_val = "_".join(suburb.replace(" ", "_") for suburb in suburbs)
                    parts.append(suburb_val)
            first = pkg_df.iloc[0]
            if diam_col and diam_col in pkg_df.columns:
                try:
                    diam_val = str(int(float(first[diam_col]))).strip()
                except (ValueError, TypeError):
                    diam_val = str(first[diam_col]).strip()
                if diam_val:
                    parts.append(f"D{diam_val}")

            filename = "_".join(parts) + ".csv"
            csv_filenames.append(filename)

            out_df = pkg_df.copy()
            out_df["Diameters (mm)"] = package_diameters_text(pkg_df)
            max_distance = package_max_distance_m(pkg_df)
            out_df["Max Distance Between Pipes (km)"] = format_distance_km(max_distance)
            drop_cols = [c for c in out_df.columns if c.startswith("_")]
            out_df = out_df.drop(columns=drop_cols, errors="ignore")
            keep = [c for c in OUTPUT_KEEP_COLS if c in out_df.columns]
            out_df = out_df[keep]

            if ASSET_LEN_COL in out_df.columns:
                out_df[ASSET_LEN_COL] = pd.to_numeric(out_df[ASSET_LEN_COL], errors="coerce").apply(
                    lambda x: f"{x:,.1f} m" if pd.notna(x) else ""
                )
            if "pipe_cost" in out_df.columns:
                out_df["pipe_cost"] = pd.to_numeric(out_df["pipe_cost"], errors="coerce").apply(
                    lambda x: f"${x:,.2f}" if pd.notna(x) else ""
                )

            totals = {col: "" for col in out_df.columns}
            totals["package_id"] = "Totals"
            if ASSET_LEN_COL in out_df.columns:
                raw_lengths = pd.to_numeric(
                    out_df[ASSET_LEN_COL].astype(str).str.replace(r"[^\d.\-]", "", regex=True),
                    errors="coerce",
                )
                totals[ASSET_LEN_COL] = f"{raw_lengths.sum():,.2f} m"
            if "pipe_cost" in out_df.columns:
                raw_cost = pd.to_numeric(
                    out_df["pipe_cost"].astype(str).str.replace(r"[$,]", "", regex=True),
                    errors="coerce",
                ).sum()
                totals["pipe_cost"] = f"${raw_cost:,.2f}"
            if ASSET_ID_COL in out_df.columns:
                totals["Number of Pipes"] = len(out_df)
                out_df["Number of Pipes"] = ""
            totals["Diameters (mm)"] = package_diameters_text(pkg_df)
            totals["Max Distance Between Pipes (km)"] = format_distance_km(max_distance)
            out_df = pd.concat([out_df, pd.DataFrame([totals])], ignore_index=True)
            z.writestr(filename, out_df.to_csv(index=False).encode("utf-8"))

        lines = ["PACKAGE SUMMARY", "=" * 60]
        for pkg_id in get_package_order(df):
            pkg_df = df[df["package_id"] == pkg_id]
            if pkg_df.empty:
                continue
            lines.append("")
            lines.append(f"Package:    {pkg_id}")
            if effective_suburb_col and effective_suburb_col in pkg_df.columns:
                suburbs = ", ".join(sorted(pkg_df[effective_suburb_col].astype(str).unique()))
                lines.append(f"Suburb:     {suburbs}")
            lines.append(f"Diameters:  {package_diameters_text(pkg_df)}")
            lines.append(f"Pipes:      {len(pkg_df)}")
            if ASSET_LEN_COL in pkg_df.columns:
                lines.append(f"Length:     {pkg_df[ASSET_LEN_COL].sum():,.1f}m")
            max_distance = package_max_distance_m(pkg_df)
            if max_distance is not None:
                lines.append(f"Max Dist:   {format_distance_km(max_distance)}")
            if "pipe_cost" in pkg_df.columns:
                lines.append(f"Cost:       ${pkg_df['pipe_cost'].sum():,.0f}")
            lines.append("-" * 60)
        z.writestr("package_summary.txt", "\n".join(lines).encode("utf-8"))

        filelist_lines = ["PACKAGE FILE LIST", "=" * 60, ""] + sorted(csv_filenames)
        z.writestr("file_list.txt", "\n".join(filelist_lines).encode("utf-8"))

        green = "0A7A3B"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor=green)

        wb_summary = openpyxl.Workbook()
        ws = wb_summary.active
        ws.title = "Package Summary"
        headers = ["Package ID", "Suburb", "Diameters (mm)", "Max Distance Between Pipes (km)", "Pipes", "Length (m)", "Cost ($)", "Asset IDs"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for pkg_id in get_package_order(df):
            pkg_df = df[df["package_id"] == pkg_id]
            if pkg_df.empty:
                continue
            suburb = ", ".join(sorted(pkg_df[effective_suburb_col].astype(str).unique())) if effective_suburb_col and effective_suburb_col in pkg_df.columns else ""
            asset_ids = ", ".join(pkg_df[ASSET_ID_COL].dropna().astype(str).str.replace(r"\.0$", "", regex=True).tolist()) if ASSET_ID_COL in pkg_df.columns else ""
            ws.append([
                pkg_id,
                suburb,
                package_diameters_text(pkg_df),
                format_distance_km(package_max_distance_m(pkg_df)),
                len(pkg_df),
                f"{round(pkg_df[ASSET_LEN_COL].sum(), 1):,.1f} m" if ASSET_LEN_COL in pkg_df.columns else "",
                round(pkg_df["pipe_cost"].sum(), 0) if "pipe_cost" in pkg_df.columns else "",
                asset_ids,
            ])
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
        summary_xlsx = io.BytesIO()
        wb_summary.save(summary_xlsx)
        z.writestr("package_summary.xlsx", summary_xlsx.getvalue())

        wb_files = openpyxl.Workbook()
        ws2 = wb_files.active
        ws2.title = "File List"
        cell = ws2.cell(row=1, column=1, value="Filename")
        cell.font = header_font
        cell.fill = header_fill
        for row_idx, name in enumerate(sorted(csv_filenames), 2):
            ws2.cell(row=row_idx, column=1, value=name)
        ws2.column_dimensions["A"].width = 80
        filelist_xlsx = io.BytesIO()
        wb_files.save(filelist_xlsx)
        z.writestr("file_list.xlsx", filelist_xlsx.getvalue())

    buffer.seek(0)
    return buffer


def _sample_package_assets(df: pd.DataFrame, suburb_col: str | None) -> list[dict[str, Any]]:
    if df.empty or "package_id" not in df.columns:
        return []

    first_package_id = str(sorted(df["package_id"].astype(str).unique())[0])
    package_df = df[df["package_id"].astype(str) == first_package_id].copy()
    sample_cols = [
        c
        for c in ["package_id", ASSET_ID_COL, suburb_col, ASSET_DIAM_COL, ASSET_LEN_COL, "pipe_cost"]
        if c and c in package_df.columns
    ]
    rows: list[dict[str, Any]] = []
    for _, row in package_df[sample_cols].head(12).iterrows():
        item = {col: _value_or_none(row[col]) for col in sample_cols}
        if "pipe_cost" in item and item["pipe_cost"] is not None:
            item["pipe_cost"] = round(float(item["pipe_cost"]), 2)
        rows.append(item)
    return rows


def _build_map_assets(df: pd.DataFrame, suburb_col: str | None, limit: int = 2500) -> list[dict[str, Any]]:
    if df.empty:
        return []

    working = df.copy()
    for coord_col in ["XMid", "YMid"]:
        if coord_col not in working.columns:
            return []
        working[coord_col] = pd.to_numeric(working[coord_col], errors="coerce")

    working = working.dropna(subset=["XMid", "YMid"])
    if working.empty:
        return []

    rows: list[dict[str, Any]] = []
    sample = working.head(limit)
    cols = [
        c
        for c in [
            "package_id",
            ASSET_ID_COL,
            suburb_col,
            ASSET_DIAM_COL,
            ASSET_LEN_COL,
            CONDITION_COL,
            "pipe_cost",
            "XMid",
            "YMid",
        ]
        if c and c in sample.columns
    ]

    for _, row in sample[cols].iterrows():
        item = {col: _value_or_none(row[col]) for col in cols}
        if item.get("pipe_cost") is not None:
            item["pipe_cost"] = round(float(item["pipe_cost"]), 2)
        rows.append(item)

    return rows


class SplitPreviewRequest(BaseModel):
    relining_mode: str = "Condition 7 and 8"


class GenerateReliningPackagesRequest(BaseModel):
    relining_mode: str = "Condition 7 and 8"
    package_method: str = "value"
    max_package_value: float = 500000.0
    pipes_per_package: int = 25
    group_by_suburb: bool = True
    cost_mode: str = "median"
    contractor: str | None = None
    packaging_mode: str | None = None
    grouping_method: str = "Suburb (no coordinates)"
    size_groups: list[list[int]] = []
    traffic_control: bool = True
    project_initiation: bool = True
    topup_mode: str = "No top up"


def _normalize_size_groups(size_groups: list[list[int]]) -> list[list[int]]:
    seen: set[int] = set()
    resolved: list[list[int]] = []
    for group in size_groups or []:
        current: list[int] = []
        for diameter in group:
            try:
                value = int(diameter)
            except Exception:
                continue
            if value in seen:
                continue
            seen.add(value)
            current.append(value)
        if current:
            resolved.append(current)
    return resolved


def _build_group_columns(
    df: pd.DataFrame,
    grouping_method: str,
    suburb_col: str | None,
) -> tuple[pd.DataFrame, list[str], str | None]:
    working = df.copy()
    group_cols: list[str] = []
    active_grouping = grouping_method or "Suburb (no coordinates)"

    if active_grouping != "Suburb (no coordinates)":
        digits = "".join(ch for ch in active_grouping if ch.isdigit())
        if digits:
            radius_m = int(digits)
            working = assign_spatial_clusters(
                working,
                radius_m=radius_m,
                x_col=X_MID_COL,
                y_col=Y_MID_COL,
                suburb_col=suburb_col or "Asset Suburb",
            )
            group_cols.append("_cluster")
            return working, group_cols, "_cluster"

    if suburb_col and suburb_col in working.columns:
        group_cols.append(suburb_col)
    return working, group_cols, suburb_col


def _cost_relining_with_options(
    df: pd.DataFrame,
    rates: pd.DataFrame,
    cost_mode: str,
    contractor: str | None,
    traffic_control: bool,
    project_initiation: bool,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    if df.empty:
        empty = df.copy()
        return empty, empty

    working = normalize_asset_columns(df.copy())
    working[ASSET_ID_COL] = working[ASSET_ID_COL].astype(str)
    working["rate_per_m"] = working.apply(
        lambda row: get_relining_rate(
            row.get(ASSET_DIAM_COL),
            row.get(ASSET_LEN_COL),
            rates,
            mode="vendor" if cost_mode == "contractor" else cost_mode,
            vendor=contractor if cost_mode == "contractor" else None,
        ),
        axis=1,
    )
    lengths = pd.to_numeric(working[ASSET_LEN_COL], errors="coerce") if ASSET_LEN_COL in working.columns else pd.Series(dtype=float)
    working["pipe_cost"] = pd.to_numeric(working["rate_per_m"], errors="coerce") * lengths
    uncosted = working[working["pipe_cost"].isna()].copy()
    working = working[working["pipe_cost"].notna()].copy()

    if not working.empty:
        if traffic_control:
            working["pipe_cost"] = working["pipe_cost"] + 2000.0
        if project_initiation:
            working["pipe_cost"] = working["pipe_cost"] * 1.15

    return working, uncosted


def _build_topup_pool(
    source_df: pd.DataFrame,
    existing_df: pd.DataFrame,
    rates: pd.DataFrame,
    cost_mode: str,
    contractor: str | None,
    traffic_control: bool,
    project_initiation: bool,
    grouping_method: str,
    suburb_col: str | None,
) -> pd.DataFrame:
    if source_df is None or source_df.empty:
        return pd.DataFrame()

    pool, _ = _cost_relining_with_options(
        source_df,
        rates=rates,
        cost_mode=cost_mode,
        contractor=contractor,
        traffic_control=traffic_control,
        project_initiation=project_initiation,
    )
    if pool.empty:
        return pool

    pool, pool_suburb = ensure_suburb_column(pool, resolve_suburb_column(pool))
    pool, _, effective_suburb = _build_group_columns(pool, grouping_method, pool_suburb)
    if ASSET_DIAM_COL in pool.columns:
        pool["_size_group"] = pool[ASSET_DIAM_COL].apply(
            lambda d: get_package_group_label(coerce_diameter_value(d)) if coerce_diameter_value(d) is not None else "unknown"
        )

    if existing_df is not None and not existing_df.empty:
        existing_ids = set(existing_df[ASSET_ID_COL].astype(str))
        pool = pool[~pool[ASSET_ID_COL].astype(str).isin(existing_ids)].copy()

    if effective_suburb and effective_suburb not in pool.columns and suburb_col and suburb_col in pool.columns:
        effective_suburb = suburb_col

    return pool


def _is_island_pipe(row: pd.Series, suburb_col: str | None = None) -> bool:
    candidate_cols = [
        suburb_col,
        "Asset Suburb",
        "Location - Suburb",
        "Pipe_Start_Address",
        "Formatted_Address",
        "Physical_Location",
        "Location - Street/Road Name",
    ]
    seen: set[str] = set()
    for col in candidate_cols:
        if not col or col in seen or col not in row.index:
            continue
        seen.add(col)
        text = str(row.get(col) or "").upper()
        if "SCOTLAND ISLAND" in text:
            return True
        if " ISLAND" in text or text.startswith("ISLAND "):
            return True
    return False


def _force_island_single_pipe_packages(
    packaged: pd.DataFrame,
    prefix: str,
    suburb_col: str | None = None,
) -> pd.DataFrame:
    if packaged is None or packaged.empty or "package_id" not in packaged.columns:
        return packaged

    result = packaged.copy()
    island_mask = result.apply(lambda row: _is_island_pipe(row, suburb_col=suburb_col), axis=1)
    if not island_mask.any():
        return result

    existing_numbers = []
    for package_id in result["package_id"].dropna().astype(str).unique():
        try:
            existing_numbers.append(int(package_id.split("_")[-1]))
        except ValueError:
            continue
    next_number = max(existing_numbers, default=0) + 1

    for idx in result[island_mask].index:
        result.at[idx, "package_id"] = f"{prefix}{next_number:03d}"
        next_number += 1

    return result


def _generate_relining_package_result(
    payload: GenerateReliningPackagesRequest,
) -> tuple[dict[str, Any], pd.DataFrame, pd.DataFrame, str | None]:
    split = _split_assets(payload.relining_mode)
    relining_df = split["relining"]
    suburb_col = split["_suburb_col"]
    assets_df = split["_assets_df"]
    rates = load_relining_rates_cached()
    raw_cond = extract_condition_scores(assets_df)

    packaging_mode = (payload.packaging_mode or "").strip() or (
        "Pipes per package" if (payload.package_method or "").strip().lower() == "count" else "Max package value"
    )
    grouping_method = payload.grouping_method or "Suburb (no coordinates)"
    size_groups = _normalize_size_groups(payload.size_groups)

    costed, uncosted = _cost_relining_with_options(
        relining_df,
        rates=rates,
        cost_mode=(payload.cost_mode or "median").strip().lower(),
        contractor=payload.contractor,
        traffic_control=payload.traffic_control,
        project_initiation=payload.project_initiation,
    )

    if costed.empty:
        response = {
            "ok": True,
            "relining_mode": payload.relining_mode,
            "package_method": (payload.package_method or "value").strip().lower(),
            "packaging_mode": packaging_mode,
            "group_by_suburb": bool(payload.group_by_suburb),
            "costed_pipe_count": 0,
            "uncosted_pipe_count": int(len(uncosted)),
            "package_count": 0,
            "total_length_m": 0.0,
            "total_cost": 0.0,
            "packages": [],
            "sample_assets": [],
            "map_assets": [],
            "base_package_count": 0,
        }
        return response, pd.DataFrame(), pd.DataFrame(), suburb_col

    costed, suburb_col = ensure_suburb_column(costed, suburb_col)
    costed, group_cols, effective_group_col = _build_group_columns(costed, grouping_method, suburb_col)

    method = (payload.package_method or "value").strip().lower()
    if packaging_mode == "Pipes per package" or method == "count":
        packaged = split_into_packages_by_count(
            costed,
            pipes_per_package=max(int(payload.pipes_per_package), 1),
            group_cols=group_cols,
            prefix="RLN_",
        )
        packaged_base = packaged.copy()
        method = "count"
    else:
        packaged = split_into_value_packages_twop(
            costed,
            max_package_value=max(float(payload.max_package_value), 1.0),
            suburb_col=effective_group_col,
            prefix="RLN_",
            cost_col="pipe_cost",
            adjacent_fill=(payload.topup_mode != "No top up"),
            custom_groups=size_groups if size_groups else None,
            group_col=effective_group_col if effective_group_col == "_cluster" else None,
        )
        packaged_base = packaged.copy()
        method = "value"

        if payload.topup_mode != "No top up":
            packaged_ids = set(packaged[ASSET_ID_COL].astype(str)) if not packaged.empty else set()
            same_condition_unpackaged = costed[
                ~costed[ASSET_ID_COL].astype(str).isin(packaged_ids)
            ].copy()
            same_condition_pool = _build_topup_pool(
                same_condition_unpackaged,
                existing_df=costed,
                rates=rates,
                cost_mode=(payload.cost_mode or "median").strip().lower(),
                contractor=payload.contractor,
                traffic_control=payload.traffic_control,
                project_initiation=payload.project_initiation,
                grouping_method=grouping_method,
                suburb_col=suburb_col,
            )

            if payload.topup_mode == "Same condition only" and not same_condition_pool.empty:
                packaged = topup_packages(
                    packaged=packaged,
                    topup_pool=same_condition_pool,
                    max_package_value=max(float(payload.max_package_value), 1.0),
                    suburb_col=effective_group_col,
                )
            elif payload.topup_mode == "Same condition first, lower condition if needed":
                if not same_condition_pool.empty:
                    packaged = topup_packages(
                        packaged=packaged,
                        topup_pool=same_condition_pool,
                        max_package_value=max(float(payload.max_package_value), 1.0),
                        suburb_col=effective_group_col,
                    )

                if payload.relining_mode == "Condition 8 only":
                    lower_source_df = assets_df[raw_cond == 7].copy()
                elif payload.relining_mode == "Condition 7 only":
                    lower_source_df = assets_df[raw_cond == 8].copy()
                else:
                    lower_source_df = pd.DataFrame()

                lower_condition_pool = _build_topup_pool(
                    lower_source_df,
                    existing_df=costed,
                    rates=rates,
                    cost_mode=(payload.cost_mode or "median").strip().lower(),
                    contractor=payload.contractor,
                    traffic_control=payload.traffic_control,
                    project_initiation=payload.project_initiation,
                    grouping_method=grouping_method,
                    suburb_col=suburb_col,
                )
                if not lower_condition_pool.empty:
                    packaged = topup_packages(
                        packaged=packaged,
                        topup_pool=lower_condition_pool,
                        max_package_value=max(float(payload.max_package_value), 1.0),
                        suburb_col=effective_group_col,
                    )

    packaged = _force_island_single_pipe_packages(packaged, prefix="RLN_", suburb_col=suburb_col)
    packaged_base = _force_island_single_pipe_packages(packaged_base, prefix="RLN_", suburb_col=suburb_col)

    packaged = add_package_priority(packaged)
    packaged = reindex_packages_by_priority(packaged, prefix="RLN_")
    packaged = add_package_priority(packaged)
    packaged = sort_packages_by_priority(packaged)

    packaged_base = add_package_priority(packaged_base)
    packaged_base = reindex_packages_by_priority(packaged_base, prefix="RLN_")
    packaged_base = add_package_priority(packaged_base)
    packaged_base = sort_packages_by_priority(packaged_base)

    package_count = int(packaged["package_id"].nunique()) if not packaged.empty and "package_id" in packaged.columns else 0
    total_cost = round(float(pd.to_numeric(packaged["pipe_cost"], errors="coerce").fillna(0).sum()), 2) if not packaged.empty and "pipe_cost" in packaged.columns else 0.0
    total_length = round(float(pd.to_numeric(packaged[ASSET_LEN_COL], errors="coerce").fillna(0).sum()), 1) if not packaged.empty and ASSET_LEN_COL in packaged.columns else 0.0

    response = {
        "ok": True,
        "relining_mode": payload.relining_mode,
        "package_method": method,
        "packaging_mode": packaging_mode,
        "grouping_method": grouping_method,
        "group_by_suburb": bool(group_cols),
        "costed_pipe_count": int(len(costed)),
        "uncosted_pipe_count": int(len(uncosted)),
        "package_count": package_count,
        "total_length_m": total_length,
        "total_cost": total_cost,
        "packages": _package_summary_rows(packaged, suburb_col),
        "sample_assets": _sample_package_assets(packaged, suburb_col),
        "map_assets": _build_map_assets(packaged, suburb_col),
        "base_package_count": int(packaged_base["package_id"].nunique()) if not packaged_base.empty else 0,
    }
    return response, packaged, packaged_base, suburb_col


app = FastAPI(title="Stormwater Packaging API", version="0.1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=get_allowed_origins(),
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

WEB_DIR = Path(__file__).resolve().parent / "web"
if WEB_DIR.is_dir():
    app.mount("/static", StaticFiles(directory=WEB_DIR), name="static")


@app.get("/")
def root() -> FileResponse:
    return FileResponse(WEB_DIR / "exb.html")


@app.get("/app")
def web_app() -> FileResponse:
    return FileResponse(WEB_DIR / "exb.html")


@app.get("/shell")
def shell_app() -> FileResponse:
    return FileResponse(WEB_DIR / "full.html")


@app.get("/lite")
def lite_app() -> FileResponse:
    return FileResponse(WEB_DIR / "index.html")


@app.get("/web/config")
def web_config() -> dict[str, Any]:
    return {
        "ok": True,
        "streamlit_url": get_streamlit_url(),
        "lite_url": "/lite",
    }


@app.get("/health")
def health() -> dict[str, Any]:
    assets_df = load_assets()
    relining_rates_path = _data_dir() / "Panel_Rates.xlsx"
    outputs_path = _outputs_dir()
    return {
        "ok": True,
        "service": "stormwater-packaging-api",
        "assets_loaded": int(len(assets_df)),
        "assets_path": str(get_assets_path()),
        "rates_path": str(relining_rates_path),
        "outputs_path": str(outputs_path),
        "has_coordinates": all(
            col in assets_df.columns for col in ["XMid", "YMid"]
        ),
    }


@app.get("/packaging/config")
def packaging_config() -> dict[str, Any]:
    assets_df = load_assets()
    suburb_col = resolve_suburb_column(assets_df)
    normalized = normalize_asset_columns(assets_df.copy())
    cond = (
        pd.to_numeric(
            normalized[CONDITION_COL].astype(str).str.extract(r"(\d+)")[0],
            errors="coerce",
        )
        if CONDITION_COL in normalized.columns
        else pd.Series(dtype=float)
    )
    diameters = []
    if ASSET_DIAM_COL in normalized.columns:
        diameters = sorted(
            {
                int(value)
                for value in normalized.loc[cond.isin([7, 8]), ASSET_DIAM_COL]
                .dropna()
                .apply(coerce_diameter_value)
                .dropna()
                .tolist()
            }
        )
        if not diameters:
            diameters = sorted(
                {
                    int(value)
                    for value in normalized[ASSET_DIAM_COL]
                    .dropna()
                    .apply(coerce_diameter_value)
                    .dropna()
                    .tolist()
                }
            )
    return {
        "ok": True,
        "assets_path": str(get_assets_path()),
        "asset_count": int(len(assets_df)),
        "suburb_column": suburb_col,
        "data_path": str(_data_dir()),
        "outputs_path": str(_outputs_dir()),
        "proximity_options_m": PROXIMITY_OPTIONS,
        "relining_modes": [
            "Condition 8 only",
            "Condition 7 only",
            "Condition 7 and 8",
        ],
        "cost_modes": ["median", "lowest", "contractor"],
        "packaging_modes": ["Max package value", "Pipes per package"],
        "grouping_methods": ["Suburb (no coordinates)"] + [f"{p}m" for p in PROXIMITY_OPTIONS],
        "topup_modes": [
            "No top up",
            "Same condition only",
            "Same condition first, lower condition if needed",
        ],
        "available_relining_diameters": diameters,
        "has_coordinates": all(col in assets_df.columns for col in [X_MID_COL, Y_MID_COL]),
    }


@app.get("/packaging/contractors")
def packaging_contractors() -> dict[str, Any]:
    rates = load_relining_rates_cached()
    return {
        "ok": True,
        "contractors": get_contractors(rates),
    }


@app.post("/packaging/split-streams")
def packaging_split_streams(payload: SplitPreviewRequest) -> dict[str, Any]:
    split = _split_assets(payload.relining_mode)
    assets_df = split["_assets_df"]
    suburb_col = split["_suburb_col"]

    return {
        "ok": True,
        "relining_mode": payload.relining_mode,
        "asset_count": int(len(assets_df)),
        "suburb_column": suburb_col,
        "data_path": str(_data_dir()),
        "outputs_path": str(_outputs_dir()),
        "streams": [
            _stream_summary("Relining Preview", split["relining"], suburb_col),
            _stream_summary(
                "Relining Default (Condition 8)", split["relining_cond_8_default"], suburb_col
            ),
            _stream_summary("Reconstruction", split["reconstruction"], suburb_col),
            _stream_summary("Amplification", split["amplification"], suburb_col),
        ],
        "map_assets": {
            "relining": _build_map_assets(split["relining"], suburb_col),
            "reconstruction": _build_map_assets(split["reconstruction"], suburb_col),
            "amplification": _build_map_assets(split["amplification"], suburb_col),
        },
    }


@app.post("/packaging/generate-relining-packages")
def packaging_generate_relining_packages(
    payload: GenerateReliningPackagesRequest,
) -> dict[str, Any]:
    response, _, _, _ = _generate_relining_package_result(payload)
    return response


@app.post("/packaging/export-relining-packages")
def export_relining_packages(
    payload: GenerateReliningPackagesRequest,
) -> Response:
    _, packaged, _, suburb_col = _generate_relining_package_result(payload)
    zip_buffer = create_zip_from_packages(
        packaged,
        suburb_col=suburb_col,
        diam_col=ASSET_DIAM_COL if not packaged.empty and ASSET_DIAM_COL in packaged.columns else None,
    )
    return Response(
        content=zip_buffer.getvalue(),
        media_type="application/zip",
        headers={"Content-Disposition": 'attachment; filename="relining_packages.zip"'},
    )



